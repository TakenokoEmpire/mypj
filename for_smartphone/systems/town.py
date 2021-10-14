from typing import List, Tuple, Optional
import random
import openpyxl
from . import battle


class Town(battle.Core):
    """街"""

    def __init__(self):
        super().__init__()
        # battle.Core()

    def main(self):

        while True:
            print("メニュー")
            print("1:ステータスチェック")
            print("2:スキル")
            print("3:装備")
            print("4:アイテム")
            print("5:ショップ")
            print("6:クエスト")
            print("9:セーブしてタイトルに戻る")
            choice = input("何する？->")
            if choice == "1":
                self.town_status()
            elif choice == "2":
                self.town_skill()
            elif choice == "3":
                self.town_equip()
            elif choice == "4":
                self.town_item()
            elif choice == "5":
                self.town_shop()
            elif choice == "6":
                self.town_quest()
            elif choice == "7":
                self.town_test()
            elif choice == "9":
                self.save()
                exit()

    def print_equip(self):
        for i in range(self.vlookup(self.mysheet, "equip_qty", 2)):
            print(("{} {} :{} (hp:+{}, atk:+{})").format(i, self.equip_position[i], self.vhlookup(self.mysheet, self.equip_position[i], 1, "name", self.equip_index_rownum), self.vhlookup(
                self.mysheet, self.equip_position[i], 1, "hp", self.equip_index_rownum), self.vhlookup(self.mysheet, self.equip_position[i], 1, "atk", self.equip_index_rownum)))

    def town_status(self):
        # battle.Battle("town")
        self.status_checker()
        self.print_equip()
        self.print_status()

    def town_equip(self):
        # battle.Battle("town")
        # 装備の情報更新
        self.equip_checker()
        equip_list = []
        # 今の装備を表示
        self.print_equip()
        # 装備変更
        s_pos = self.hindex(self.book["装備"], "position")
        s_name = self.hindex(self.book["装備"], "name")
        s_qty = self.hindex(self.book["装備"], "qty")
        choice_position = int(
            input("どの装備を変更しますか？[0~{}]->".format(self.equip_qty-1)))
        for i in range(500):
            # 部位が一致する装備を一覧表示するためにリストに追加
            # 所持数が空白だとエラー起きる
            if self.book["装備"].cell(row=i+1, column=s_pos).value == self.equip_position[choice_position] and self.book["装備"].cell(row=i+1, column=s_qty).value >= 1:
                equip_list.append(self.book["装備"].cell(
                    row=i+1, column=s_name).value)
        for i in range(len(equip_list)):
            print("{}:{}".format(i, equip_list[i]))
        choice_equips = int(
            input("どれと交換しますか？[0~{}]->".format(len(equip_list)-1)))
        self.mysheet[self.vhindex(self.mysheet, self.equip_position[choice_position],
                                  1, "name", self.equip_index_rownum, "excel")] = equip_list[choice_equips]

        # 変更した状態を表示
        # battle.Battle("town")
        self.status_checker()
        self.print_equip()
        self.print_status()
        print("装備を変更しました。")

        # self.save()
        # self.book.close
        # self.book = openpyxl.load_workbook('systems/base.xlsx', data_only=True)

    def town_equip1(self):
        """"現在の全身装備をリストで返す"""
        # 装備の情報更新
        self.equip_checker()
        # 今の装備を表示
        position_choice_list = []
        position_detail_list = []
        for i in range(self.vlookup(self.mysheet, "equip_qty", 2)):
            choice = (("{} :{}").format(self.equip_position[i], self.vhlookup(
                self.mysheet, self.equip_position[i], 1, "name", self.equip_index_rownum)))
            detail = ((("hp:+{}, atk: +{}").format(self.vhlookup(self.mysheet, self.equip_position[i], 1, "hp", self.equip_index_rownum), self.vhlookup(
                self.mysheet, self.equip_position[i], 1, "atk", self.equip_index_rownum))))
            position_choice_list.append(choice)
            position_detail_list.append(choice+" : "+detail)
        return position_choice_list, position_detail_list

    def town_equip2(self, choice_position):
        """手持ちの装備のうち、指定された部位のものをリストで返す"""
        # 装備変更
        equip_choice_list = []
        s_pos = self.hindex(self.book["装備"], "position")
        s_name = self.hindex(self.book["装備"], "name")
        s_qty = self.hindex(self.book["装備"], "qty")
        for i in range(500):
            # 部位が一致する装備を一覧表示するためにリストに追加
            # 所持数が空白だとエラー起きる
            if self.book["装備"].cell(row=i+1, column=s_pos).value == self.equip_position[choice_position] and self.book["装備"].cell(row=i+1, column=s_qty).value >= 1:
                equip_choice_list.append(self.book["装備"].cell(
                    row=i+1, column=s_name).value)
        return equip_choice_list

    def town_equip3(self, choice_position, equip_list, choice_equips):
        """指定された部位の装備品リストの中から、指定された装備を現在装備と交換する"""
        self.mysheet[self.vhindex_super(self.mysheet, self.equip_position[choice_position],
                                  1, "name", self.equip_index_rownum, "excel")] = equip_list[choice_equips]
        print(self.mysheet[self.vhindex(self.mysheet, self.equip_position[choice_position],
                                        1, "name", self.equip_index_rownum, "excel")].value)
        print("装備が変更されました！")
        # battle.Battle("town")
        self.status_checker()
        self.print_equip()
        self.print_status()
        print(self.book["プレイヤーステータス"]["B4"].value)

    def town_equip3_new(self, selected_position,selected_id):
        """指定された部位の装備品リストの中から、指定された装備を現在装備と交換する"""
        self.mysheet[self.vhindex_super(self.mysheet, selected_position,
                                  1, "id", "position", "excel","off")] = int(selected_id)
        print("装備が変更されました！")
        # battle.Battle("town")
        # self.equip_checker_new()
        # self.print_equip()
        # self.print_status()
        # print(self.book["プレイヤーステータス"]["B4"].value)

    def town_test(self):
        self.mysheet["G6"].value += 1
        print("value:{}".format(self.mysheet["G6"].value))
