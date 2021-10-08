import openpyxl
import pprint
import random
import math
# from . import battle
# class Initialize(battle.Battle):
"""めんどくさいからinitはそのままコピーした
    """


class Initialize():
    def __init__(self):
        pass
        """コンストラクタ
        lv,hp,atk,skill[0]~[2],exp,moneyはリストとなっており、[0]がプレイヤー、[1]が敵
        """
        self.basic_status_index = ["lv", "hp", "atk"]
        self.mobname = "スライム"
        self.book = openpyxl.load_workbook('systems/base.xlsx', data_only=True)
        self.mobsheet = self.book[self.mobname]
        self.mysheet = self.book["プレイヤーステータス"]
        self.vs_sheet = [self.mysheet, self.mobsheet]
        self.lv = [0, 0]
        self.hp = [0, 0]
        self.atk = [0, 0]
        self.skill = [[0, 0] for j in range(3)]
        self.exp = [0, 0]
        self.money = [0, 0]
        for num, sheet in enumerate(self.vs_sheet):
            self.lv[num] = self.vlookup(sheet, "lv", 2)
            self.hp[num] = self.vlookup(sheet, "hp", 2)
            self.atk[num] = self.vlookup(sheet, "atk", 2)
            for j in range(3):
                self.skill[j][num] = self.vlookup(
                    sheet, "skill{}".format(j), 2)
            self.exp[num] = self.vlookup(sheet, "exp", 2)
            self.money[num] = self.vlookup(sheet, "money", 2)
        self.droplist = []

    def confirm(self):
        confirm = input("これまでのプレイデータを削除します。本当によろしいですか？[Y/n] ->")
        if confirm == "y" or confirm == "Y":
            input("削除されたデータは戻りません。本当によろしいですか？[Y/n] ->")
            if confirm == "y" or confirm == "Y":
                self.all_zero()
                print("データが削除されました。")

    def all_zero(self):
        """戦闘後の処理を参考にした
        ステータスの更新
        経験値の更新
        金の更新
        アイテムの更新
        以上の情報をエクセルファイルに更新・保存
        """
        # ステータス更新
        init_lv = 0
        init_hp = 100
        init_atk = 10
        init_stats = [init_lv, init_hp, init_atk]
        init_val = [1, 1, 0, 0]
        # init_dict = {"total":1,"raw":1,"equip":0,"temporal":0}
        # equipの欄はいじらなくても、equipを削除するようにすれば勝手に0になりそう。
        for i, stat in enumerate(self.basic_status_index):
            for j in range(4):
                self.mysheet.cell(row=self.x_index(
                    self.mysheet, stat), column=j + 2, value=init_stats[i]*init_val[j])

        # 経験値、金の更新
        self.mysheet.cell(row=self.x_index(
            self.mysheet, "exp"), column=2, value=0)
        self.mysheet.cell(row=self.x_index(
            self.mysheet, "money"), column=2, value=0)
        # ドロップアイテム数は1のみ対応。
        # 解説：ゲットしたアイテムがまだ道具箱に1個もない→新しくインデックスを追加し個数を1増やす(実際に行う順序は、個数を1にしてからインデックス追加)
        # 　　　そうでない（すでにある）→個数を1増やす
        for i in range(39):
            self.book["道具箱"].cell(row=i+2, column=1, value="empty")
            self.book["道具箱"].cell(row=i+2, column=2, value=0)

        """開いてる途中だとエラー出るよ"""
        try:
            self.book.save("systems/base.xlsx")
        except PermissionError:
            print("エクセルファイルを閉じてください")  # これ戦闘前に欲しい
            exit()

    def vlookup(self, sheet, index, column_order):
        """excelのvlookup関数を再現。
        「完全一致」のみ。
        """
        for i in range(500):
            output = 0
            if str(sheet.cell(row=i + 1, column=1).value) == index:
                output = sheet.cell(row=i+1, column=column_order).value
                return output
        return "False"

    def hlookup(self, sheet, index, raw_order):
        """excelのmatch関数を再現"""
        for i in range(100):
            output = 0
            if str(sheet.cell(row=i + 1, column=1).value) == index:
                output = sheet.cell(row=raw_order, column=i+1).value
                return output
        return "False"

    def x_index(self, sheet, index):
        """excelのindex関数もどきを再現
        該当する文字がある行数を返す（先頭列に限る）"""
        for i in range(500):
            output = 0
            if str(sheet.cell(row=i + 1, column=1).value) == index:
                return i + 1
        return "False"
