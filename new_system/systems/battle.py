import openpyxl
import pprint
import random
import math
import copy
from systems import judge_manual
import try_1003
import try_1009
import send_recieve
import time
import unicodedata
from typing import List, Tuple

# from systems import testtry

"""よくやるミス
self忘れ
=と==
関数の()
return忘れ（Noneが帰ってきたときや、Nonetypeとの足し算ができないと言われたとき注意）
gameplay以外のファイルを実行（ipmortがおかしいときは大抵これ）
過剰にリスト化（text must be a unicode or bytesのとき）
変数のstr化（text must be a unicode or bytes）
Excel検索文字列がstrでない（text must be a unicode or bytes）
"""

"""
やること
runにまとめる
通常戦：（単発のhitblowで終わらない場合）→とき終わったら、再度答えを生成
ボス戦（単発のhitblowで終わらない場合）→どちらかが終わったら、アルゴリズム再起動
"""

"""
やりたいこと
special mode→クリティカル攻撃
"""

"""メモ
ゲーム序盤は2回、中盤は4回、終盤は7回ほどで次のダンジョンに行けるようゲームバランス調整
（慣れれば慣れるほど、解くのが早くなって所要時間が短縮される）
"""


# drun = SendReceive(room_id=room)

# room_id:Optional[int]=None,
# self.room_id = room_id

class Function():
    """計算で使う関数たち
    ゲーム的要素を一切含まない、ただの計算用の関数"""

    def __init__(self):
        self.switch_judge_box = {}

    def vlookup(self, sheet, index, column_order):
        """excelのvlookup関数を再現。
        先頭行・列からスタートで固定"""
        for i in range(500):
            output = 0
            if str(sheet.cell(row=i + 1, column=1).value) == index:
                output = sheet.cell(row=i+1, column=column_order).value
                return output
        return False

    def vlookup_plus(self, sheet, index, column_order, left_index_position):
        """excelのvlookup関数を再現。
        スタート位置をずらせる
        """
        for i in range(500):
            output = 0
            if str(sheet.cell(row=i + 1, column=left_index_position).value) == index:
                output = sheet.cell(
                    row=i+1, column=left_index_position+column_order-1).value
                return output
        return False

    def hlookup(self, sheet, index, row_order):
        """excelのhlookup関数を再現
        先頭行・列からスタートで固定"""
        for i in range(100):
            output = 0
            if str(sheet.cell(row=1, column=i+1).value) == index:
                output = sheet.cell(row=row_order, column=i+1).value
                return output
        return False

    def hlookup_plus(self, sheet, index, row_order, top_index_position):
        """excelのhlookup関数を再現。
        スタート位置をずらせる
        """
        for i in range(100):
            output = 0
            if str(sheet.cell(row=top_index_position, column=i+1).value) == index:
                output = sheet.cell(
                    row=row_order+top_index_position-1, column=i+1).value
                return output
        return False

    def vhlookup(self, sheet, left_index, left_index_position, top_index, top_index_position,minus_search = "off", multi_find = "off" ):
        """excelの、vlookup関数とhlookup関数を組み合わせた関数
        indexとmatchを組み合わせて作るアレ
        minus_search="on"にすると、top_indexを探す際に、left_index_positionの行よりも左側の行からも探す（列も同様）
        """
        if multi_find == "on":
            return_box = []
            if minus_search=="on":
                for i in range(500):
                    if str(sheet.cell(row=i+1, column=left_index_position).value) == left_index:
                        for j in range(100):
                            if str(sheet.cell(row=top_index_position, column=j+1).value) == top_index:
                                return_box.append(sheet.cell(row=i+1, column=j+1).value)
                return return_box
            else:
                for i in range(500):
                    if str(sheet.cell(row=i + top_index_position, column=left_index_position).value) == left_index:
                        for j in range(100):
                            if str(sheet.cell(row=top_index_position, column=j+left_index_position).value) == top_index:
                                return_box.append(sheet.cell(row=i+top_index_position, column=j+left_index_position).value)
                return return_box
        else:
            if minus_search=="on":
                for i in range(500):
                    if str(sheet.cell(row=i+1, column=left_index_position).value) == left_index:
                        for j in range(100):
                            if str(sheet.cell(row=top_index_position, column=j+1).value) == top_index:
                                return sheet.cell(row=i+1, column=j+1).value
                return False
            else:
                for i in range(500):
                    if str(sheet.cell(row=i + top_index_position, column=left_index_position).value) == left_index:
                        for j in range(100):
                            if str(sheet.cell(row=top_index_position, column=j+left_index_position).value) == top_index:
                                return sheet.cell(row=i+top_index_position, column=j+left_index_position).value
                return False
            

    def vhlookup_super(self, sheet, left_index, left_index_position_or_topname, top_index, top_index_position_or_leftname,minus_search = "off", multi_find = "off"):
        """vhlookupについて、position指定をその行・列の先頭セルの名前でも行えるようになった。
        第3、第5引数について、intを入力した場合は従来通り、strを入力した場合その文字で探すようになる。
        minus_search="on"にすると、top_indexを探す際に、left_index_positionの行よりも左側の行からも探す（列も同様）"""
        if type(left_index_position_or_topname) == str:
            left_index_position = self.hindex(sheet,left_index_position_or_topname)
        else:
            left_index_position = left_index_position_or_topname
        if type(top_index_position_or_leftname) == str:
            top_index_position = self.vindex(sheet,top_index_position_or_leftname)
        else:
            top_index_position = top_index_position_or_leftname
        return self.vhlookup(sheet, left_index, left_index_position, top_index, top_index_position,minus_search,multi_find)

    def vindex(self, sheet, index):
        """excelのindex関数もどきを再現
        該当する文字がある行番号を返す（先頭列に限る）"""
        for i in range(500):
            output = 0
            if str(sheet.cell(row=i + 1, column=1).value) == index:
                return i + 1
        return False

    def vindex_plus(self, sheet, index, left_index_position):
        """excelのindex関数もどきを再現
        該当する文字がある行番号を返す"""
        for i in range(500):
            output = 0
            if str(sheet.cell(row=i + 1, column=left_index_position).value) == index:
                return i + 1
        return False

    def hindex(self, sheet, index):
        """excelのindex関数もどきを再現
        該当する文字がある列番号を返す（先頭行に限る）"""
        for i in range(100):
            output = 0
            if str(sheet.cell(row=1, column=i+1).value) == index:
                return i + 1
        return False

    def hindex_plus(self, sheet, index, top_index_number):
        """excelのindex関数もどきを再現
        該当する文字がある列番号を返す"""
        for i in range(100):
            output = 0
            if str(sheet.cell(row=top_index_number, column=i+1).value) == index:
                return i + 1
        return False

    def vhindex(self, sheet, left_index, left_index_position, top_index, top_index_position, return_type,minus_search = "off"):
        """excelのindex関数もどきを再現
        縦横両方向に対応し、さらに先頭行（列）以外にも対応
        return_type:"excel"→セル番号形式
                    その他→row,column
        minus_search="on"にすると、top_indexを探す際に、left_index_positionの行よりも左側の行からも探す（列も同様）
        """
        if minus_search == "on":
            for i in range(500):
                output = 0
                if str(sheet.cell(row=i + 1, column=left_index_position).value) == left_index:
                    # print("a")
                    for j in range(100):
                        #     print(sheet.cell(row=j + 1, column=1).value)
                        #     print(top_index)
                        if str(sheet.cell(row=top_index_position, column=j+1).value) == top_index:
                            if return_type == "excel":
                                return openpyxl.utils.get_column_letter(j + 1)+str(i + 1)
                            else:
                                return i + 1, j + 1
        else:
            for i in range(500):
                output = 0
                if str(sheet.cell(row=i + top_index_position, column=left_index_position).value) == left_index:
                    # print("a")
                    for j in range(100):
                        #     print(sheet.cell(row=j + 1, column=1).value)
                        #     print(top_index)
                        if str(sheet.cell(row=top_index_position, column=j+left_index_position).value) == top_index:
                            if return_type == "excel":
                                return openpyxl.utils.get_column_letter(j + left_index_position)+str(i + top_index_position)
                            else:
                                return i + top_index_position, j + left_index_position
            return False

    def vhindex_super(self, sheet, left_index, left_index_position_or_topname, top_index, top_index_position_or_leftname,return_type,minus_search = "off"):
        """動作未確認。
        vhindexについて、position指定をその行・列の先頭セルの名前でも行えるようになった。
        第3、第5引数について、intを入力した場合は従来通り、strを入力した場合その文字で探すようになる。
        minus_search="on"にすると、top_indexを探す際に、left_index_positionの行よりも左側の行からも探す（列も同様）"""
        if type(left_index_position_or_topname) == str:
            left_index_position = self.hindex(sheet,left_index_position_or_topname)
        else:
            left_index_position = left_index_position_or_topname
        if type(top_index_position_or_leftname) == str:
            top_index_position = self.vindex(sheet,top_index_position_or_leftname)
        else:
            top_index_position = top_index_position_or_leftname
        return self.vhindex(sheet, left_index, left_index_position, top_index, top_index_position,return_type,minus_search)


    def zero_false(self, x):
        if x == 0:
            return False
        else:
            return True

    def switch_judge(self, target_var_name_str, target_var_value, switch_value):
        """
        前回実行時targetの値がswitch_valueと一致せず、かつ今回一致した場合にTrueを返す、
        GUI再生時を想定している
        複数の関数を対象にしてもバグらないように、判定対象の関数名(str)を引数として取る（識別できれば関数名はなんでもよい）"""
        # ↓のやり方すれば、「選択肢」モジュールをもうちょっと簡単にできそう
        if (target_var_name_str in self.switch_judge_box) == False:
            self.switch_judge_box[target_var_name_str] = [None, None]
        if switch_value == None:
            print("switch_value should not be None")
            exit()
        self.switch_judge_box[target_var_name_str][0] = target_var_value
        if self.switch_judge_box[target_var_name_str][0] != self.switch_judge_box[target_var_name_str][1] and self.switch_judge_box[target_var_name_str][0] == switch_value:
            print("switch_value:TRUE")
            self.switch_judge_box[target_var_name_str][1] = self.switch_judge_box[target_var_name_str][0]
            self.switch_judge_box[target_var_name_str][0] = None
            return True
        else:
            self.switch_judge_box[target_var_name_str][1] = self.switch_judge_box[target_var_name_str][0]
            self.switch_judge_box[target_var_name_str][0] = None
            return False

    def text_length(self, text):
        """半角換算での文字列の長さを返す"""
        count = 0
        for c in text:
            if unicodedata.east_asian_width(c) in 'FWA':
                count += 2
            else:
                count += 1
        return count

    def line_break(self, text, line_length):
        """全角半角を考慮した自動改行ツール
        text:改行対象文字列
        line_length:1行の長さ（半角換算）"""
        text_box = []
        while len(text) > 0:
            for i in range(len(text)+1):
                if self.text_length(text[:i]) > line_length-1:
                    text_box.append(text[:i])
                    text = text[i:]
                    break
                elif self.text_length(text) <= line_length-1:
                    text_box.append(text)
                    text = ""
                    break
        return text_box


    def if_return(self,target,threshold,return_value_when_false):
        """targetがthresholdと一致したらreturn_value_when_falseを返す
        そうでない場合はtargetをそのまま返す"""
        if target == threshold:
            return return_value_when_false
        else:
            return target

    def randex(self,min_val,max_val,exponent=1):
        rand = random.random()
        rand_exponent = rand ** exponent

        return min_val + rand_exponent * (max_val - min_val)

    def rand_judge(self,threshold_prob):
        rand = random.random()
        if rand < threshold_prob:
            return True
        else:
            return False

class Core(Function):
    """
    変数：複数の場面で使用されるもの
    関数：複数の場面で使用されるもの
    Battleクラスから、ダンジョン情報が必要ないものでかつ他でも使うものをこちらに移管した。
    街での作業、ダンジョンでの戦闘を完了すると、再起動するように設定されている（本当に？）。"""

    def __init__(self):
        super().__init__()
        print("CORE")
        # エクセルファイルの呼び出し
        try:
            self.book = openpyxl.load_workbook(
                'systems/base.xlsx', data_only=True)
        except:
            self.book = openpyxl.load_workbook(
                'C:/Users/wolke/git1009/new_system/systems/base.xlsx', data_only=True)
        self.mysheet = self.book["プレイヤーステータス"]

        # 課金関連
        self.diamond = self.vlookup(self.mysheet, "diamond", 2)
        # 装備関連
        self.equip_position = ["右手", "左手", "鎧", "靴", "装飾品"]
        self.equip_index_rownum = self.vindex(self.mysheet, "position")
        self.equip_sheetsoubi_name_columnnum = self.hindex(
            self.book["装備"], "name")
        self.equip_qty = self.vlookup(self.mysheet, "equip_qty", 2)
        # ステータス関連
        """
        lv,hp,atk,skill1~3,exp,moneyは、プレイヤーの値を返すint（スキルはstr）にする。（スキルも[0]~[2]ではなく、それぞれのstrとなっている）
        敵のステータスは、全て頭にe_をつける。
        対戦時はリストとなっており、[0]がプレイヤーの値、[1]が敵の値。
        """
        # 便利な一覧表
        self.basic_status_index = ["lv", "hp", "atk"]
        self.all_status_index = ["lv", "hp", "atk", "exp", "money",
                                 "skill1", "skill2", "skill3"]
        self.all_plus_attr_status_index = ["lv", "hp", "atk", "exp", "money",
                                 "skill1", "skill2", "skill3","water_attr","plant_attr","dark_attr","elect_attr"]
        # self.basic_status_index_var = [self.lv, self.hp, self.atk]
        # self.all_status_index_var = [self.lv, self.hp, self.atk, self.exp, self.money, self.skill1, self.skill2, self.skill3]
        # 属性関連
        self.attr_dict = {"dark":"闇","water":"水","plant":"木","elect":"雷","all":"全","dark_attr":"闇","water_attr":"水","plant_attr":"木","elect_attr":"雷","all_attr":"全"}
        self.attr_list4 = ["water_attr","plant_attr","dark_attr","elect_attr"]
        self.attr_list5 = ["water_attr","plant_attr","dark_attr","elect_attr","all_attr"]
        # ステータス取得
        self.status_checker()  # ステータス取得前にステータスを更新しておく

    # def equip_index_of_position(self, position):

    #     return equip_index_list, equip_sepc_list

    def status_checker(self):
        self.attribute_update()
        self.equip_checker_new()
        update_index = self.basic_status_index + self.attr_list4
        for stat in update_index:
            self.mysheet[self.vhindex(self.mysheet, stat, 1, "total", 1, "excel")] = self.vhlookup(self.mysheet, stat, 1, "raw", 1) + self.vhlookup(self.mysheet, stat, 1, "equip", 1)
        status_box = []
        for num in range(len(self.all_plus_attr_status_index)):
            status_box.append(self.vlookup(self.mysheet, self.all_plus_attr_status_index[num], 2))
        
        # ステータス定義
        self.lv = status_box[0]
        self.hp = status_box[1]
        self.atk = status_box[2]
        self.exp = status_box[3]
        self.money = status_box[4]
        self.skill1 = status_box[5]
        self.skill2 = status_box[6]
        self.skill3 = status_box[7]
        self.attr_power = status_box[8:12]
        

    def equip_checker(self):
        # 装備欄に登録された名前から、hp,atkの値を「装備」ブックから参照し記録する
        update_stat = ["hp", "atk"]
        sum_stat = [0, 0]
        for i in range(self.equip_qty):
            equip_name = self.vhlookup(
                self.mysheet, self.equip_position[i], 1, "name", self.equip_index_rownum)
            for stat in update_stat:
                self.mysheet[self.vhindex(self.mysheet, self.equip_position[i], 1, stat, self.equip_index_rownum, "excel")] = self.vhlookup(
                    self.book["装備"], equip_name, self.equip_sheetsoubi_name_columnnum, stat, 1)
        # 装備欄に記録されたhp,atkの補正値をステータスに反映する(補正値の合計sum_statを計算し、該当する位置に貼り付ける)
        for j in range(len(update_stat)):
            for i in range(self.equip_qty):
                sum_stat[j] += self.vhlookup(self.mysheet, self.equip_position[i],
                                             1, update_stat[j], self.equip_index_rownum)
            self.mysheet[self.vhindex(
                self.mysheet, update_stat[j], 1, "equip", 1, "excel")] = sum_stat[j]

    def equip_checker_new(self):
        # 装備欄に登録された名前から、update_statリストの値を「装備」ブックから参照し記録する
        # update_statとsum_statの大きさは合わせること
        # 原因不明のバグのため、プレイヤーステータスシートの装備欄は更新せず、IDのみを情報として使う。
        update_stat = ["hp","atk","water_attr","plant_attr","dark_attr","elect_attr"]
        sum_stat = [0, 0,0,0,0,0]
        for j,stat in enumerate(update_stat):
            for i in range(len(self.equip_position)):
                equip_id = self.vhlookup_super(self.mysheet, self.equip_position[i], 1, "id", "position","off")
                sum_stat[j] += self.id_equip_info(equip_id,stat)
            self.mysheet[self.vhindex(
                self.mysheet, stat, 1, "equip", 1, "excel")] = sum_stat[j]

        # for i in range(len(self.equip_position)):
        #     equip_id = self.vhlookup_super(self.mysheet, self.equip_position[i], 1, "id", "position","off")
        #     for stat in update_stat:
        #         self.mysheet[self.vhindex_super(self.mysheet, self.equip_position[i], 1, stat, "position", "excel","off")] = self.id_equip_info(self.vhlookup_super(self.mysheet, self.equip_position[i], 1, stat, "position","off"),stat)
        # # 装備欄に記録されたhp,atkの補正値をステータスに反映する(補正値の合計sum_statを計算し、該当する位置に貼り付ける)
        # for j in range(len(update_stat)):
        #     for i in range(len(self.equip_position)):
        #         print(self.vhlookup_super(self.mysheet, self.equip_position[i], 1, update_stat[j], "position","off"))
        #         print(self.vhlookup_super(self.mysheet, self.equip_position[i], 1, "equip_atk", "position","off"))
                
        #         print(self.equip_position[i])
        #         print(update_stat[j])
        #         sum_stat[j] += self.vhlookup_super(self.mysheet, self.equip_position[i],
        #                                      1, update_stat[j], "position","off")
        # #ステータスもこの関数で更新できるように変更した。→できてなかった
        # for stat in self.basic_status_index:
        #     self.mysheet[self.vhindex(self.mysheet, stat, 1, "total", 1, "excel")] = self.vhlookup(self.mysheet, stat, 1, "raw", 1) + self.vhlookup(self.mysheet, stat, 1, "equip", 1)

    def print_status(self):
        print("my status")
        print("lv,hp,atk,exp")
        print([self.lv, self.hp, self.atk, self.exp])

    def gacha(self, sheet, gachaname):
        rand = random.random()
        top_index_position = self.vindex(sheet, gachaname)-1
        gacha_qty = self.vlookup(sheet, gachaname, 2)
        for i in range(gacha_qty):
            if rand < float(self.hlookup_plus(sheet, "sum_prob", i+2, top_index_position)):
                selected_item = self.hlookup_plus(
                    sheet, "content", i+2, top_index_position)
                item_rarity = self.hlookup_plus(
                    sheet, "rarity", i+2, top_index_position)
                return selected_item, item_rarity

    def get_item(self, item: str, add_qty: int = 1):
        if self.vindex_plus(self.book["装備"], item, 3) != False:
            self.get_equip(item,add_qty)
        elif self.vindex_plus(self.book["アイテム箱"], item, 2) != False:
            self.get_nonequip_item(item,add_qty)
        else:
        # 登録されてないアイテムをゲットした場合、道具箱に投げる
            print("登録されていないアイテムを取得しました。当該アイテムは道具箱に登録されます。")
            if self.vindex(self.book["道具箱"], item) == False:
                self.book["道具箱"].cell(row=self.vindex(self.book["道具箱"], "empty"), column=2, value=self.vlookup(self.book["道具箱"], "empty", 2)+1)
                self.book["道具箱"].cell(row=self.vindex(self.book["道具箱"], "empty"), column=1, value=item)
            else:
                self.book["道具箱"].cell(row=self.vindex(self.book["道具箱"], item),column=2, value=self.vlookup(self.book["道具箱"], item, 2)+1)

    def get_equip(self, item: str, add_qty: int = 1):
        rand = random.random()
        item_pos = self.vindex_plus(self.book["装備"], item, 3)
        qty_pos = self.hindex(self.book["装備"], "qty")  # qtyではなく"qty"なので注意
        # c_value = self.vlookup(self.book["装備"], item, 3)
        equip_body_position = self.vhlookup(self.book["装備"], item, 3, "position_copy", 1)
        equip_hp = self.vhlookup(self.book["装備"], item, 3, "hp", 1)
        equip_atk = self.vhlookup(self.book["装備"], item, 3, "atk", 1)
        equip_def = self.vhlookup(self.book["装備"], item, 3, "def", 1)
        # 「装備」タブへの処理
        self.book["装備"].cell(row=item_pos,column=qty_pos, value=self.vlookup_plus(self.book["装備"], item, 2, 3)+add_qty)
        # 「個別装備情報」タブへの処理
        equip_rownum = self.vindex_plus(self.book["装備個別情報"], "empty", 3)
        for i in range(5):
            if rand < float(self.hlookup(self.book["装備個別情報"], "sum_prob", i+2)):
                slot = i+1
                break
        self.book["装備個別情報"].cell(row=equip_rownum, column=self.hindex(self.book["装備個別情報"],"position"), value=equip_body_position)
        self.book["装備個別情報"].cell(row=equip_rownum, column=self.hindex(self.book["装備個別情報"],"name"), value=item)
        self.book["装備個別情報"].cell(row=equip_rownum, column=self.hindex(self.book["装備個別情報"],"hp"), value=equip_hp)
        self.book["装備個別情報"].cell(row=equip_rownum, column=self.hindex(self.book["装備個別情報"],"atk"), value=equip_atk)
        self.book["装備個別情報"].cell(row=equip_rownum, column=self.hindex(self.book["装備個別情報"],"def"), value=equip_def)
        self.book["装備個別情報"].cell(row=equip_rownum, column=self.hindex(self.book["装備個別情報"],"slot_qty"), value=slot)
        print("{}:{}を入手した！slot:{}".format(equip_body_position,item,slot))

    def get_nonequip_item(self, item: str, add_qty: int = 1):
        """item:アイテム名、add_qty:取得した数"""
        item_pos = self.vindex_plus(self.book["アイテム箱"], item, 2)
        qty_pos = self.hindex(self.book["アイテム箱"], "qty")  # qtyではなく"qty"なので注意
        self.book["アイテム箱"].cell(row=item_pos,column=qty_pos, value=self.item_info(item,"qty")+add_qty)

    def buy_item(self, item: str, value: int, diamond: int = 0, add_qty: int = 1):
        """item:アイテム名,value:価格,diamond:リアルマネー、add_qty:取得した数"""
        if self.check_minus(value,diamond) == False:
            return False
        self.money -= value
        self.diamond -= diamond
        self.get_item(item, add_qty)
        self.mysheet.cell(row=self.vindex(self.mysheet, "money"),
                          column=2, value=self.money)
        self.mysheet.cell(row=self.vindex(self.mysheet, "diamond"),
                          column=2, value=self.diamond)

    def buy_diamond(self, diamond_qty):
        self.diamond += diamond_qty
        self.mysheet.cell(row=self.vindex(self.mysheet, "diamond"),
                          column=2, value=self.diamond)

    def check_minus(self,money_qty_to_use,diamond_qty_to_use = 0):
        if money_qty_to_use > self.money:
            return False
        elif diamond_qty_to_use > self.diamond:
            return False
        else:
            return True

    def item_info(self,item_name,info_type):
        return self.vhlookup_super(self.book["アイテム箱"],item_name,"name",info_type,1,"on")

    def item_multi_info(self,item_name,info_type_list:List[str]):
        box = []
        for info_type in info_type_list:
            box.append(self.item_info(item_name,info_type))
        return box

    def use_item(self,item_name,qty = 1):
        qty_minus = -qty
        self.get_item(item_name,qty_minus)

    def use_multi_item(self,item_list):
        """使用数は1で固定"""
        for item_name in item_list:
            if item_name == "":
                pass
            else:
                self.use_item(item_name,1)

    def item_type_list(self,item_type,info_type,only_own_mode="off"):
        """指定したitem_typeに属する全てのアイテムを返す
        only_owa_modeをonにすると、1個以上持っているもののみを返す"""
        if only_own_mode != "on":
            return self.vhlookup_super(self.book["アイテム箱"],item_type,"type",info_type,1,"on","on")

        else:
            box1 = self.vhlookup_super(self.book["アイテム箱"],item_type,"type",info_type,1,"on","on")
            box2 = self.vhlookup_super(self.book["アイテム箱"],item_type,"type","qty",1,"on","on")
            count = 0
            while True:
                try:
                    if box2[count] == 0:
                        box1.pop(count)
                        box2.pop(count)
                    else:
                        count += 1
                except IndexError:
                    break
            return box1
        # else:
        #     box1 = self.vhlookup_super(self.book["アイテム箱"],item_type,"type",info_type,1,"on","on")
        #     box2 = self.vhlookup_super(self.book["アイテム箱"],item_type,"type","qty",1,"on","on")
        #     box12 = list(zip(box1,box2))
        #     for num, qty in enumerate(box2):
        #         if qty == 0:

        #     return box1

    def item_type_list_multi_info(self,item_type,info_type_list,only_own_mode="off"):
        box = []
        for info_type in info_type_list:
            box.append(self.item_type_list(item_type,info_type,only_own_mode))
        return box

    def equip_list_position(self,position,info_type):
        """指定された部位の、保有装備の一覧表を返す"""
        return self.vhlookup_super(self.book["装備個別情報"],position,"position",info_type,1,"on","on")

    def id_equip_info(self,equip_id,info_type):
        return self.vhlookup_super(self.book["装備個別情報"],str(equip_id),"id",info_type,1,"on")

    def id_equip_multi_info(self,equip_id,info_type_list:List[str]):
        box = []
        for info_type in info_type_list:
            box.append(self.id_equip_info(equip_id,info_type))
        return box

    def id_equip_dict(self,equip_id):
        """idが一致する装備のすべての情報をdictで返す。"""
        key = ["id","position","name","atk","slot_qty","slot_empty","water_attr","plant_attr","dark_attr","elect_attr"]
        key.extend(list(("slot{}_attr".format(i+1)for i in range(5))))
        key.extend(list(("slot{}_val".format(i+1)for i in range(5))))
        val = self.id_equip_multi_info(equip_id,key)
        equip_info_dict = dict(zip(key,val))
        # print(equip_info_dict)
        return equip_info_dict

    def current_equip_list(self, info_type):
        box = []
        for pos in self.equip_position:
            box.append(self.vhlookup_super(self.mysheet,pos,1,info_type,"position","off"))
        return box

    def equip_val_update(self,equip_id,info_type,update_val):
        # print(equip_id,info_type,update_val)
        # if type(update_val) in [int,float]:
        #     update_val = str(update_val)
        # self.book["装備個別情報"][self.vhindex_super(self.book["装備個別情報"],equip_id,1,info_type,1,"Excel")] = update_val
        equip_rownum = self.vindex_plus(self.book["装備個別情報"], str(equip_id), 1)
        equip_columnnum=self.hindex_plus(self.book["装備個別情報"],info_type,1)
        self.book["装備個別情報"].cell(row=equip_rownum, column=equip_columnnum, value=update_val)

    def equip_multi_val_update(self,equip_id,info_type_list,update_val_list):
        # print(equip_id,info_type_list,update_val_list)
        # box = []
        for i in range(len(info_type_list)):
            self.equip_val_update(equip_id,info_type_list[i],update_val_list[i])
        # return box

    def compound(self,equip_id,slot_num:int,material,catalyst_red = "",catalyst_blue = "",catalyst_green = ""):
        """装備に様々な能力を付与する成長合成を行う。
        リリース段階では、武器に属性値を付与するもののみを実装する予定
        現段階では、合成値＝【攻撃力】＊合成係数　となっているため、今後防具にも適用する際は注意"""
        dic = self.id_equip_dict(equip_id)
        # if dic["slot{}_attr".format(slot_num)] != None:
        #     return "slot not empty"
        attr, compound_value_ratio, min_val_ratio,max_val_ratio,rand,rand_exponent = self.compound_judge(material,catalyst_red,catalyst_blue,catalyst_green)
        compound_value = round(compound_value_ratio * dic["atk"] / 100)
        print("attr,rand,rand_exponent,compound_value_ratio,compound_value")
        print([attr,rand,rand_exponent,compound_value_ratio,compound_value])
        self.use_multi_item([material,catalyst_red,catalyst_blue,catalyst_green])
        self.equip_multi_val_update(equip_id,["slot{}_attr".format(slot_num),"slot{}_val".format(slot_num)],[attr,compound_value])
        self.attribute_update(equip_id)
        return attr,compound_value

    def compound_judge(self,material,catalyst_red = "",catalyst_blue = "",catalyst_green = ""):
        """同種類の触媒を同時に送ることは禁止（特に緑系統）
        合成値= 最小値 + (最大値-最小値)*乱数^x"""
        #xは、デフォルトでは2。平均値は0.33。低い値が出やすいようになっている。
        #緑系統触媒の「合成値○○倍」は、xの値が小さくなることによってこの平均値がどれほど上昇するかを表している。
        #最大の1.5倍の触媒を使うと、x=1で平均値が0.5。これでやっとフラット。
        #重課金させるためには、0.2~0.3にしてもいいかも。

        material_info = self.item_multi_info(material,["effect_type","effect_val","effect_val_max"])
        if catalyst_red != "":
            max_up_ratio = self.item_info(catalyst_red,"effect_val")
        else:
            max_up_ratio = 1
        if catalyst_blue != "":
            min_up_ratio = self.item_info(catalyst_blue,"effect_val")
        else:
            min_up_ratio = 1
        if catalyst_green != "":
            exponent_up_ratio = self.item_info(catalyst_green,"effect_val")
        else:
            exponent_up_ratio = 1

        attr = material_info[0]
        min_val = material_info[1] * min_up_ratio
        max_val = material_info[2] * max_up_ratio
        exponent = 2 * exponent_up_ratio

        rand = random.random()
        rand_exponent = rand ** exponent

        compound_value_ratio = min_val + rand_exponent * (max_val - min_val)

        return attr,compound_value_ratio,min_val,max_val,rand,rand_exponent

    def compound_min_max(self,equip_id,slot_num:int,material,catalyst_red = "",catalyst_blue = "",catalyst_green = ""):
        """合成値判定前に、ゲーム内表示用に触媒込みの最大最小を返す"""
        dic = self.id_equip_dict(equip_id)
        # if dic["slot{}_attr".format(slot_num)] != None:
        #     return "slot not empty"
        attr, compound_value_ratio, min_val_ratio,max_val_ratio,rand,rand_exponent = self.compound_judge(material,catalyst_red,catalyst_blue,catalyst_green)
        min_val = round(min_val_ratio * dic["atk"] / 100)
        max_val = round(max_val_ratio * dic["atk"] / 100)
        return min_val,max_val

    def catalyst_judge(self,target_catalyst,current_catalyst_red,current_catalyst_blue,current_catalyst_green):
        catalyst_red = current_catalyst_red
        catalyst_blue = current_catalyst_blue
        catalyst_green = current_catalyst_green
        if self.item_info(target_catalyst,"effect_type") == "compound_max_up":
            catalyst_red = target_catalyst
        elif self.item_info(target_catalyst,"effect_type") == "compound_min_up":
            catalyst_blue = target_catalyst
        elif self.item_info(target_catalyst,"effect_type") == "compound_luck_up":
            catalyst_green = target_catalyst
        # catalyst_red_detail = "{}:最大値上昇+{}%".format(catalyst_red,self.if_return(self.item_info(catalyst_red,"effect_val_max"),False,0))
        # catalyst_blue_detail = "{}:最小値上昇+{}%".format(catalyst_blue,self.if_return(self.item_info(catalyst_blue,"effect_val_max"),False,0))
        # catalyst_green_detail = "{}:ランダム数値上昇+{}%".format(catalyst_green,self.if_return(self.item_info(catalyst_green,"effect_val_max"),False,0))
        return catalyst_red,catalyst_blue,catalyst_green

    def attribute_update(self,equip_id = "all"):
        """それぞれのスロットの属性値を計算し、それぞれの属性の値を書き込む。
        さらに、スロット数を補正する。
        装備IDを指定しない場合、全装備を計算する。"""
        if equip_id == "all":
            for i in range(self.get_equip_qty()):
                current_id = i+1
                dic = self.id_equip_dict(current_id)
                attr_list = ["water_attr","plant_attr","dark_attr","elect_attr"]
                for attr in attr_list:
                    attr_val = 0
                    for slot in range(5):
                        if dic["slot{}_attr".format(slot+1)] == attr:
                            attr_val += dic["slot{}_val".format(slot+1)]
                    self.equip_val_update(current_id,"{}".format(attr),attr_val)
                used_slot_count = 0
                for slot in range(5):
                    if self.id_equip_info(current_id,"slot{}_val".format(slot+1)) != None:
                        used_slot_count += 1
                self.equip_val_update(current_id,"slot_empty",self.id_equip_info(current_id,"slot_qty")-used_slot_count)

        else:
            dic = self.id_equip_dict(equip_id)
            attr_list = ["water_attr","plant_attr","dark_attr","elect_attr"]
            for attr in attr_list:
                attr_val = 0
                for slot in range(5):
                    if dic["slot{}_attr".format(slot+1)] == attr:
                        attr_val += dic["slot{}_val".format(slot+1)]
                self.equip_val_update(equip_id,"{}".format(attr),attr_val)
            used_slot_count = 0
            for slot in range(5):
                if self.id_equip_info(equip_id,"slot{}_val".format(slot+1)) != None:
                    used_slot_count += 1
            self.equip_val_update(equip_id,"slot_empty",self.id_equip_info(equip_id,"slot_qty")-used_slot_count)

    def get_equip_qty(self):
        for i in range(500):
            if self.id_equip_info(i+1,"name") == "empty":
                return i
        
    def decompound(self,equip_id,slot_num):
        print(equip_id,slot_num)
        self.use_item("魔女の秘薬")
        #Noneを書き込もうとするとなぜか書き込まれないので力技で
        # for row in self.book["装備個別情報"].iter_rows(min_row=equip_id+1, min_col=13+slot_num*2, max_row=1, max_col=2):
        #     for cell in row:
        #         cell.value = None
        self.book["装備個別情報"].cell(row=equip_id+1,column=slot_num*2+13).value=None
        self.book["装備個別情報"].cell(row=equip_id+1,column=slot_num*2+14).value=None
        self.attribute_update(equip_id)

    def save(self):
        """開いてる途中だとエラー出るよ"""
        try:
            self.book.save("systems/base.xlsx")
        except FileNotFoundError:
            self.book.save(
                'C:/Users/wolke/git1009/new_system/systems/base.xlsx')
        except PermissionError:
            print("エクセルファイルを閉じてください")  # これ戦闘前に欲しい
            exit()



class Battle(Core):
    """ダメージを入力し、その結果を出力するように子クラスを設計中
    主にエクセルファイルとやり取りする
    """

    def __init__(self, place: str = "battle"):
        super().__init__()
        """
    戦闘関連の情報処理全般を扱う。
    ダンジョン以外でも使うものは、Coreへ移管した。
    """

        # GUIに対応
        # ダンジョンに依存するステータス等は、この段階まで待つ必要があるため、
        # 進入ダンジョンが決定する前(initで強制的に実行される内容)と後でself設定を分割する。

        # 戦闘時だけでなく、Townからもここにアクセスする。
        # その場合は、通常ダンジョンのダンジョン番号1にアクセスしたものとして扱う（便宜上、敵も設定する必要がある）
        if place == "console":
            print("NOTICE: class Battle is being accessed by console")
            self.gamescene = int(input("Normal:1,Boss:2 ->"))
            self.dungeon_num = int(input("dungeon num 1~ (not 0)->"))
            self.dungeon_init()



    def dungeon_init(self):
        # ダンジョン情報と初期設定
        print("gamescene:{},dungeon_num:{}".format(
            self.gamescene, self.dungeon_num))
        dungeon_type_list = ["none", "normal", "boss"]
        self.dungeon_type = dungeon_type_list[self.gamescene]
        self.ans = ["12345", "abcde"]
        self.turn_count = [0, 0]
        self.hist = [[], []]
        self.max_ans = 16
        self.length = 5

        self.mobname = self.vlookup(
            self.choose_dungeon(), str(self.dungeon_num), 3)
        self.mobsheet = self.book[self.mobname]
        self.vs_sheet = [self.mysheet, self.mobsheet]
        print("野生の{}が現れた！".format(self.mobname))
        self.droplist = []

        # 敵ステータス取得
        status_box = []
        for num in range(len(self.all_status_index)):
            status_box.append(self.vlookup(
                self.mobsheet, self.all_status_index[num], 2))
        # ステータス定義
        self.e_lv = status_box[0]
        self.e_hp = status_box[1]
        self.e_atk = status_box[2]
        self.e_exp = status_box[3]
        self.e_money = status_box[4]
        self.e_skill1 = status_box[5]
        self.e_skill2 = status_box[6]
        self.e_skill3 = status_box[7]

        # """
    # lv,hp,atk,skill[0]~[2],exp,moneyはリストとなっており、[0]がプレイヤーの値、[1]が敵の値
    # """
    # self.basic_status_index = ["lv", "hp", "atk"]
    # self.status_checker()  # ステータス取得前にステータスを更新しておく
    # self.lv = [0, 0]
    # self.hp = [0, 0]
    # self.atk = [0, 0]
    # self.skill = [[0, 0] for j in range(3)]
    # self.exp = [0, 0]
    # self.money = [0, 0]
    # for num, sheet in enumerate(self.vs_sheet):
    #     self.lv[num] = self.vlookup(sheet, "lv", 2)
    #     self.hp[num] = self.vlookup(sheet, "hp", 2)
    #     self.atk[num] = self.vlookup(sheet, "atk", 2)
    #     for j in range(3):
    #         self.skill[j][num] = self.vlookup(
    #             sheet, "skill{}".format(j), 2)
    #     self.exp[num] = self.vlookup(sheet, "exp", 2)
    #     self.money[num] = self.vlookup(sheet, "money", 2)

    def choose_dungeon(self):
        # 小文字化させたい
        while True:
            if self.dungeon_type == "normal":
                return self.book["通常ダンジョン"]
            elif self.dungeon_type == "boss":
                # ボス戦の場合、この段階でアルゴリズムを最後まで回す。
                # なお、16進5桁以外は非対応である。
                self.autoplay()
                return self.book["ボスダンジョン"]
            else:
                self.dungeon_type = "normal"
                print("normalダンジョンに入ります。")
                return self.book["通常ダンジョン"]

    def main(self):
        # self.second_init_battle()
        self.print_status()
        print(self.hist)
        while True:
            if self.my_turn() == "finish":
                break
            if self.mob_turn() == "finish":
                break
            self.end_of_turn()

    def end_of_turn(self):
        if len(self.hist[0]) == min(self.turn_count)-1:
            self.hist[0].append({"guess": "-----", "hit": "0", "blow": "0"})
        elif len(self.hist[0]) == min(self.turn_count):
            pass
        else:
            print("error end_of_turn")
        self.print_history()
        print("残りHP{}".format(self.hp))

        # try:
        #     self.print_out_history()
        # except IndexError:
        #     self.hist[0].append({"guess": "-----", "hit": "0", "blow": "0"})
        #     self.print_out_history()
        # print("残りHP{}".format(self.hp))

    def print_enemy_status(self):
        # 敵情報と自分のステータスを分離するために関数を分離させた
        print("mob status")
        print("lv,hp,atk,exp")
        print([self.e_lv, self.e_hp, self.e_atk, self.e_exp])

    def print_history(self):
        for i in range(min(self.turn_count)):
            if self.dungeon_type == "normal":
                print("turn{}:".format(i) + str(self.hist[0][i]))
            elif self.dungeon_type == "boss":
                print("turn{}:".format(i) +
                      str(self.hist[0][i]) + str(self.hist[1][i]))

    """
    プレイヤーの行動
    """

    def my_turn(self):
        while True:
            self.turn_count[0] += 1
            act = input("0:攻撃,1:スキル,2:アイテム,3:逃げる ->")
            if act == "0":
                if self.jamming_judge("player") == "stop":
                    return "continue"
                self.e_hp -= int(self.atk * self.my_atk_ratio())
                if self.e_hp <= 0:
                    self.win()
                    return "finish"
                else:
                    return "continue"
            else:
                print("未対応です")

    def jamming_judge(self, defendant):
        """defandant: プレイヤーのターンならplayer、ボスのターンならboss"""
        if defendant == "player":
            lv_ratio = self.lv / self.e_lv
        elif defendant == "enemy" or defendant == "boss":
            lv_ratio = self.e_lv / self.lv
        else:
            print("defandant name error")
            exit()
        print("level ratio:{}".format(lv_ratio))
        if random.random() > lv_ratio and self.dungeon_type == "boss":
            print("jamming_judge:STOP")
            return "stop"
        return "go"

    def my_atk_ratio(self):
        # if input("do hit blow? [y/n] ->") == "y":
        while True:
            guess = input("input guess ->")
            if self.guess_checker(guess) == True:
                break
        runner = judge_manual.JudgeManual(
            guess, self.ans[0], self.length, self.max_ans)
        hit, blow = runner.run()
        self.hist[0].append({"guess": guess, "hit": hit, "blow": blow})
        if hit == 5:
            print("必殺の一撃！")
            return 1
        else:
            return hit * 0.1 + blow * 0.05
        #     # ショートカット用。
        # else:
        #     try:
        #         choice = float(input("my atk ratio->"))
        #     except ValueError:
        #         choice = 0
        #     return choice

    def guess_checker(self, guess):
        if len(guess) != self.length:
            print("{}文字で入力してください。".format(self.length))
            return False
        guess_list = []
        for i in range(len(guess)):
            try:
                int(guess[i], base=self.max_ans)
            except ValueError:
                print("不正な数字・文字が入力されています。各桁の最大数字は{}までです。".format(self.max_ans - 1))
                return False
            guess_list.append(guess[i])
        if len(guess_list) != len(set(guess_list)):
            print("同じ数字を複数の桁に入力することはできません。")
            return False
        return True

    """
    敵の行動
    """

    def autoplay(self):
        vs = try_1009.AutoPlay("off")
        self.hist[1] = vs.run()

    def mob_turn(self):
        self.hp -= int(self.e_atk * self.mob_atk_ratio())
        if self.hp <= 0:
            self.lose()
            return "finish"
        else:
            return "continue"

    def mob_atk_ratio(self):
        self.turn_count[1] += 1
        if self.dungeon_type == "normal":
            return 1
        # ボスの場合、アルゴリズムにhitblowを解かせる。
        # その際、内部処理としては、一度全て解かせて、全てのターンの情報を保存する。
        # プレイヤーが見える情報は、その内の一部の情報のみとする。
        if self.dungeon_type == "boss":
            #     print("----------")
            #     print(self.hist[1][self.turn_count[1] - 1])
            hit = self.hist[1][self.turn_count[1] - 1]["hit"]
            blow = self.hist[1][self.turn_count[1] - 1]["blow"]
            if hit == 3:
                print("強烈な力を感じる…")
            if hit == 5:
                print("周囲が光に包まれた！")
                return 1
            else:
                return hit * 0.05 + blow * 0.01

    """
    戦闘終了後
    """

    def win(self):
        print("敵を撃破した！")
        self.exp += self.e_exp
        self.money += self.e_money
        self.level_up_judge()
        self.drop()
        self.after_battle()

    def lose(self):
        print("You lose...")

    def level_up_judge(self):
        while True:
            if self.exp > self.vlookup(self.book["level_table"], str(self.lv + 1), 2):
                print("おめでとう！レベルが{}に上がった！".format(self.lv + 1))
                print("HPが{}上がった！".format(self.vlookup(
                    self.book["level_table"], str(self.lv + 1), 7)))
                print("攻撃が{}上がった！".format(self.vlookup(
                    self.book["level_table"], str(self.lv + 1), 8)))
                self.lv += 1
            else:
                break

    def drop(self) -> None:
        rand = random.random()
        for i in range(6):
            if rand < float(self.vlookup(self.mobsheet, "drop" + str(i + 1), 4)):
                print("{}を手に入れた！".format(self.vlookup(
                    self.mobsheet, "drop" + str(i + 1), 2)))
                self.droplist.append(self.vlookup(
                    self.mobsheet, "drop" + str(i + 1), 2))
                return(self.droplist)
        print(self.droplist)

    def after_battle(self):
        """戦闘後の処理
        ステータスの更新
        経験値の更新
        金の更新
        アイテムの更新
        以上の情報をエクセルファイルに更新・保存
        """
        # ステータス更新
        new_lv = self.lv
        new_hp = self.vlookup(self.book["level_table"], str(new_lv), 3)
        new_atk = self.vlookup(self.book["level_table"], str(new_lv), 4)
        new_stats = [new_lv, new_hp, new_atk]
        for i, stat in enumerate(self.basic_status_index):
            self.mysheet[self.vhindex(
                self.mysheet, stat, 1, "raw", 1, "excel")] = new_stats[i]
        # self.mysheet[self.xy_index(
        #     self.mysheet, "lv", 1, "raw", 1, "excel")] = new_lv
        # self.mysheet[self.xy_index(
        #     self.mysheet, "hp", 1, "raw", 1, "excel")] = new_hp
        # self.mysheet[self.xy_index(
        #     self.mysheet, "atk", 1, "raw", 1, "excel")] = new_atk
        # self.status_checker()
        # 経験値、金の更新
        self.mysheet.cell(row=self.vindex(self.mysheet, "exp"), column=2, value=self.exp)
        self.mysheet.cell(row=self.vindex(self.mysheet, "money"), column=2, value=self.money)
        #ドロップアイテムの処理
        self.get_item(self.droplist[0])
        self.save()


        # # self.mysheet.cell(row=self.vindex(self.mysheet, "diamond"),column=2, value=self.diamond)
        # # ドロップアイテム数は1のみ対応。
        # # 解説：ゲットしたアイテムがまだ道具箱に1個もない→新しくインデックスを追加し個数を1増やす(実際に行う順序は、個数を1にしてからインデックス追加)
        # # 　　　そうでない（すでにある）→個数を1増やす
        # if self.vindex(self.book["道具箱"], self.droplist[0]) == False:
        #     self.book["道具箱"].cell(row=self.vindex(
        #         self.book["道具箱"], "empty"), column=2, value=self.vlookup(self.book["道具箱"], "empty", 2)+1)
        #     self.book["道具箱"].cell(row=self.vindex(
        #         self.book["道具箱"], "empty"), column=1, value=self.droplist[0])
        # else:
        #     self.book["道具箱"].cell(row=self.vindex(self.book["道具箱"], self.droplist[0]),
        #                           column=2, value=self.vlookup(self.book["道具箱"], self.droplist[0], 2)+1)
        # print(self.vindex(self.book["道具箱"], "empty"))
        """開いてる途中だとエラー出るよ"""
        # self.save()

    # def save(self):
    #     """開いてる途中だとエラー出るよ"""
    #     try:
    #         self.book.save("systems/base.xlsx")
    #     except FileNotFoundError:
    #         self.book.save(
    #             'C:/Users/wolke/git1008/new_system/systems/base.xlsx')
    #     except PermissionError:
    #         print("エクセルファイルを閉じてください")  # これ戦闘前に欲しい
    #         exit()


# class Initialize(Battle):
#     """めんどくさいからinitはそのままコピーした
#     """

#     def __init__(self) -> None:
#         """コンストラクタ
#         lv,hp,atk,skill[0]~[2],exp,moneyはリストとなっており、[0]がプレイヤー、[1]が敵
#         """
#         self.basic_status_index = ["lv", "hp", "atk"]
#         self.mobname = "スライム"
#         self.book = openpyxl.load_workbook('systems/base.xlsx', data_only=True)
#         self.mobsheet = self.book[self.mobname]
#         self.mysheet = self.book["プレイヤーステータス"]
#         self.vs_sheet = [self.mysheet, self.mobsheet]
#         self.lv = [0, 0]
#         self.hp = [0, 0]
#         self.atk = [0, 0]
#         self.skill = [[0, 0] for j in range(3)]
#         self.exp = [0, 0]
#         self.money = [0, 0]
#         for num, sheet in enumerate(self.vs_sheet):
#             self.lv[num] = self.vlookup(sheet, "lv", 2)
#             self.hp[num] = self.vlookup(sheet, "hp", 2)
#             self.atk[num] = self.vlookup(sheet, "atk", 2)
#             for j in range(3):
#                 self.skill[j][num] = self.vlookup(
#                     sheet, "skill{}".format(j), 2)
#             self.exp[num] = self.vlookup(sheet, "exp", 2)
#             self.money[num] = self.vlookup(sheet, "money", 2)
#         self.droplist = []

#     def confirm(self):
#         confirm = input("これまでのプレイデータを削除します。本当によろしいですか？[Y/n] ->")
#         if confirm == "y" or confirm == "Y":
#             input("削除されたデータは戻りません。本当によろしいですか？[Y/n] ->")
#             if confirm == "y" or confirm == "Y":
#                 self.all_zero()
#                 print("データは削除されました m9(^Д^)")

#     def all_zero(self):
#         """戦闘後の処理を参考にした
#         ステータスの更新
#         経験値の更新
#         金の更新
#         アイテムの更新
#         以上の情報をエクセルファイルに更新・保存
#         """
#         # ステータス更新
#         init_lv = 0
#         init_hp = 100
#         init_atk = 10
#         box = [init_lv, ]
#         for stat in self.basic_status_index:
#             for i in range(4):
#                 self.mysheet.cell(row=self.vindex(
#                     self.mysheet, stat), column=i+2, value=init_lv)
#                 self.mysheet.cell(row=self.vindex(
#                     self.mysheet, stat), column=i+2, value=init_hp)
#                 self.mysheet.cell(row=self.vindex(
#                     self.mysheet, s), column=i+2, value=init_atk)
#         # 経験値、金の更新
#         self.mysheet.cell(row=self.vindex(
#             self.mysheet, "exp"), column=2, value=0)
#         self.mysheet.cell(row=self.vindex(
#             self.mysheet, "money"), column=2, value=0)
#         # ドロップアイテム数は1のみ対応。
#         # 解説：ゲットしたアイテムがまだ道具箱に1個もない→新しくインデックスを追加し個数を1増やす(実際に行う順序は、個数を1にしてからインデックス追加)
#         # 　　　そうでない（すでにある）→個数を1増やす
#         for i in range(39):
#             self.book["道具箱"].cell(row=i+2, column=1, value="empty")
#             self.book["道具箱"].cell(row=i+2, column=2, value=0)
#         self.save()
