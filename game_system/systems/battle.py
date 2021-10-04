import openpyxl
import pprint
import random
import math

"""よくやるミス
self忘れ
=と==
関数の()
"""


        # drun = SendReceive(room_id=room)
        
        # room_id:Optional[int]=None,
        # self.room_id = room_id

class Battle():
    """モンスターのステータスを得る
    主にエクセルファイルとやり取りする
    """


    def __init__(self, dungeon_num) -> None:
        """コンストラクタ
        lv,hp,atk,skill[0]~[2],exp,moneyはリストとなっており、[0]がプレイヤーの値、[1]が敵の値
        """
        
        self.dungeon_num = dungeon_num
        self.book = openpyxl.load_workbook('systems/base.xlsx', data_only=True)
        self.mobname = self.vlookup(self.book["ダンジョン"], str(self.dungeon_num), 3)
        self.mobsheet = self.book[self.mobname]
        self.mysheet = self.book["プレイヤーステータス"]
        self.vs_sheet = [self.mysheet, self.mobsheet]
        self.lv = [0,0]
        self.hp = [0,0]
        self.atk = [0, 0]
        self.skill = [[0, 0] for j in range(3)]
        self.exp = [0,0]
        self.money = [0,0]
        for num,sheet in enumerate(self.vs_sheet):
            self.lv[num] = self.vlookup(sheet,"lv",2)
            self.hp[num] = self.vlookup(sheet,"hp",2)
            self.atk[num] = self.vlookup(sheet, "atk", 2)
            for j in range(3):
                self.skill[j][num] = self.vlookup(sheet, "skill{}".format(j), 2)
            self.exp[num] = self.vlookup(sheet, "exp", 2)
            self.money[num] = self.vlookup(sheet, "money", 2)
        self.droplist = []
        print("野生の{}が現れた！".format(self.mobname))

    def main(self):
        self.status_checker()
        while True:
            if self.my_turn() == "finish":
                break
            if self.mob_turn() == "finish":
                break
            print("残りHP{}".format(self.hp))

    def my_turn(self):
        self.hp[1] -= int(self.atk[0] * self.my_attack())
        if self.hp[1] <= 0:
            self.win()
            return "finish"
        else:
            return "continue"

    def my_attack(self):
        try:
            choice = float(input("atk ratio->"))
        except ValueError:
            choice = 1
        return choice

    def mob_turn(self):
        self.hp[0] -= self.atk[1]
        if self.hp[0] <= 0:
            self.lose()
            return "finish"
        else:
            return "continue"

    def win(self):
        print("敵を撃破した！")
        self.exp[0] += self.exp[1]
        self.money[0] += self.money[1]
        self.level_up_judge()
        self.drop()
        self.after_battle()

    def lose(self):
        print("You lose...")

    def level_up_judge(self):
        while True:
            if self.exp[0] > self.vlookup(self.book["level_table"], str(self.lv[0] + 1), 2):
                print("おめでとう！レベルが{}に上がった！".format(self.lv[0] + 1))
                print("HPが{}上がった！".format(self.vlookup(self.book["level_table"], str(self.lv[0] + 1), 7)))
                print("攻撃が{}上がった！".format(self.vlookup(self.book["level_table"], str(self.lv[0] + 1), 8)))
                self.lv[0] += 1
            else:
                break

    def vlookup(self, sheet, index, column_order):
        """excelのvlookup関数を再現。
        「完全一致」のみ。
        """
        for i in range(500):
            output = 0
            if str(sheet.cell(row=i + 1, column=1).value) == index:
                output = sheet.cell(row=i+1, column=column_order).value
                return output
        return "False"

    # def hlookup(self, sheet, index, raw_order):
    #     """excelのhlookup関数を再現
    #     動作未確認"""
    #     for i in range(100):
    #         output = 0
    #         if str(sheet.cell(row=i + 1, column=1).value) == index:
    #             output = sheet.cell(row=raw_order, column=i+1).value
    #             return output
    #     return "False"

    def x_index(self, sheet, index):
        """excelのindex関数もどきを再現
        該当する文字がある行番号を返す（先頭列に限る）"""
        for i in range(500):
            output = 0
            if str(sheet.cell(row=i + 1, column=1).value) == index:
                return i + 1
        return "False"

    def status_checker(self) -> None:
        print("my status")
        print("lv,hp,atk,exp")
        print([self.lv[0],self.hp[0],self.atk[0],self.exp[0]])
        print("mob status")
        print("lv,hp,atk,exp")
        print([self.lv[1],self.hp[1],self.atk[1],self.exp[1]])

    def drop(self) -> None:
        rand = random.random()
        for i in range(6):
            if rand < float(self.vlookup(self.mobsheet, "drop" + str(i + 1), 4)):
                print("{}を手に入れた！".format(self.vlookup(self.mobsheet, "drop" + str(i + 1), 2)))
                self.droplist.append(self.vlookup(self.mobsheet, "drop" + str(i + 1), 2))
                return(self.droplist)
        print(self.droplist)

    def after_battle(self):
        """戦闘後の処理
        ステータスの更新
        経験値の更新
        金の更新
        アイテムの更新
        以上の情報をエクセルファイルに更新・保存
        """
        # ステータス更新
        new_lv = self.lv[0]
        new_hp = self.vlookup(self.book["level_table"], str(new_lv), 3)
        new_atk = self.vlookup(self.book["level_table"], str(new_lv), 4)
        self.mysheet.cell(row=self.x_index(self.mysheet, "lv"), column=2, value=new_lv)
        self.mysheet.cell(row=self.x_index(self.mysheet, "hp"), column=2, value=new_hp)
        self.mysheet.cell(row=self.x_index(self.mysheet, "atk"), column=2, value=new_atk)
        # 経験値、金の更新
        self.mysheet.cell(row=self.x_index(self.mysheet, "exp"), column=2, value=self.exp[0])
        self.mysheet.cell(row=self.x_index(self.mysheet, "money"), column=2, value=self.money[0])
        # ドロップアイテム数は1のみ対応。
        # 解説：ゲットしたアイテムがまだ道具箱に1個もない→新しくインデックスを追加し個数を1増やす(実際に行う順序は、個数を1にしてからインデックス追加)
        # 　　　そうでない（すでにある）→個数を1増やす
        if self.x_index(self.book["道具箱"], self.droplist[0]) == "False":
            self.book["道具箱"].cell(row=self.x_index(self.book["道具箱"], "empty"), column=2, value = self.vlookup(self.book["道具箱"], "empty",2)+1)
            self.book["道具箱"].cell(row=self.x_index(self.book["道具箱"], "empty"), column=1, value=self.droplist[0])
        else:
            self.book["道具箱"].cell(row=self.x_index(self.book["道具箱"], self.droplist[0]), column=2, value=self.vlookup(self.book["道具箱"], self.droplist[0],2)+1)
        print(self.x_index(self.book["道具箱"], "empty"))
        """開いてる途中だとエラー出るよ"""
        try:
            self.book.save("systems/base.xlsx")
        except PermissionError:
            print("エクセルファイルを閉じてください")  # これ戦闘前に欲しい
            exit()